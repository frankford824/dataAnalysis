from __future__ import annotations

import csv
import io
import tempfile
import zipfile
from collections import deque
from collections.abc import Iterator, Mapping
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import PurePosixPath
from typing import Any, BinaryIO, Literal, Protocol, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .parse.encoding import CsvEncodingError, detect_csv_encoding
from .parse.zip_safe import SafeZipPolicy, UnsafeZipError, inspect_zip
from .snapshot import SnapshotStore

MAX_WINDOW_RADIUS = 200
MAX_WINDOW_COLUMNS = 256
MAX_HEADER_ROWS = 20
MAX_ARCHIVE_MEMBERS = 256
MAX_ARCHIVE_MEMBER_BYTES = 512 * 1024 * 1024
MAX_ARCHIVE_TOTAL_BYTES = 2 * 1024 * 1024 * 1024
MAX_COMPRESSION_RATIO = 250.0
_COPY_CHUNK_SIZE = 1024 * 1024
_SUPPORTED_MEMBER_SUFFIXES = {".csv", ".xlsx"}


class EvidencePreviewError(ValueError):
    """The requested evidence cannot be represented as a safe bounded preview."""


class EvidencePreviewSecurityError(EvidencePreviewError):
    """The source failed an archive, path, or package safety check."""


class EvidencePreviewNotFoundError(EvidencePreviewError):
    """The requested member, sheet, row, or column does not exist."""


@dataclass(frozen=True, slots=True)
class EvidenceCell:
    value: str | None
    value_kind: Literal[
        "blank",
        "boolean",
        "date",
        "datetime",
        "error",
        "formula",
        "number",
        "text",
        "time",
    ]
    formula: str | None = None
    deterministic: bool = True


@dataclass(frozen=True, slots=True)
class EvidenceColumn:
    index: int
    label: str
    source_label: str


@dataclass(frozen=True, slots=True)
class EvidenceRow:
    source_row_number: int
    source_end_row_number: int
    data_row_number: int
    cells: tuple[EvidenceCell, ...]


@dataclass(frozen=True, slots=True)
class EvidenceWindow:
    header_row_number: int
    target_row_number: int
    target_data_row_number: int
    start_row_number: int
    end_row_number: int
    columns: tuple[EvidenceColumn, ...]
    rows: tuple[EvidenceRow, ...]
    target_column_index: int | None


@dataclass(frozen=True, slots=True)
class EvidenceSheet:
    name: str
    hidden: bool
    window: EvidenceWindow


@dataclass(frozen=True, slots=True)
class EvidenceWorkbook:
    content_sha256: str
    file_kind: Literal["csv", "xlsx"]
    member_name: str | None
    sheet_names: tuple[str, ...]
    sheet: EvidenceSheet
    read_only: bool = True
    formulas_are_deterministic: bool = False

    def to_dict(self) -> dict[str, object]:
        """Return the frontend DTO shape without exposing filesystem paths."""

        return {
            "contentSha256": self.content_sha256,
            "fileKind": self.file_kind,
            "memberName": self.member_name,
            "sheetNames": list(self.sheet_names),
            "readOnly": self.read_only,
            "formulasAreDeterministic": self.formulas_are_deterministic,
            "sheet": {
                "name": self.sheet.name,
                "hidden": self.sheet.hidden,
                "window": {
                    "headerRowNumber": self.sheet.window.header_row_number,
                    "targetRowNumber": self.sheet.window.target_row_number,
                    "targetDataRowNumber": self.sheet.window.target_data_row_number,
                    "startRowNumber": self.sheet.window.start_row_number,
                    "endRowNumber": self.sheet.window.end_row_number,
                    "columns": [
                        {
                            "index": column.index,
                            "label": column.label,
                            "sourceLabel": column.source_label,
                        }
                        for column in self.sheet.window.columns
                    ],
                    "rows": [
                        {
                            "sourceRowNumber": row.source_row_number,
                            "sourceEndRowNumber": row.source_end_row_number,
                            "dataRowNumber": row.data_row_number,
                            "cells": [
                                {
                                    "value": cell.value,
                                    "valueKind": cell.value_kind,
                                    "formula": cell.formula,
                                    "deterministic": cell.deterministic,
                                }
                                for cell in row.cells
                            ],
                        }
                        for row in self.sheet.window.rows
                    ],
                    "targetColumnIndex": self.sheet.window.target_column_index,
                },
            },
        }


@dataclass(frozen=True, slots=True)
class _SourceSelection:
    stream: BinaryIO
    file_kind: Literal["csv", "xlsx"]
    member_name: str | None


@dataclass(frozen=True, slots=True)
class _CsvRecord:
    record_number: int
    source_row_number: int
    source_end_row_number: int
    values: tuple[str, ...]


class _CsvReader(Protocol):
    line_num: int

    def __iter__(self) -> Iterator[list[str]]: ...


class EvidenceViewer:
    """Read a small, immutable evidence window from a content-addressed snapshot.

    The viewer never returns a local path and never evaluates spreadsheet
    formulas. Row and column limits are enforced server-side before a DTO is
    produced, so callers cannot accidentally send a whole workbook to a browser.
    """

    def __init__(
        self,
        snapshot_store: SnapshotStore,
        *,
        max_window_radius: int = MAX_WINDOW_RADIUS,
        max_columns: int = MAX_WINDOW_COLUMNS,
    ) -> None:
        if max_window_radius < 0 or max_window_radius > MAX_WINDOW_RADIUS:
            raise ValueError(f"max_window_radius must be between 0 and {MAX_WINDOW_RADIUS}")
        if max_columns <= 0 or max_columns > MAX_WINDOW_COLUMNS:
            raise ValueError(f"max_columns must be between 1 and {MAX_WINDOW_COLUMNS}")
        self.snapshot_store = snapshot_store
        self.max_window_radius = max_window_radius
        self.max_columns = max_columns

    def preview(
        self,
        content_sha256: str,
        *,
        member_name: str | None = None,
        sheet_name: str | None = None,
        target_row_number: int,
        window_radius: int = 20,
        max_columns: int = 80,
        target_column: str | int | None = None,
        header_row_number: int | None = None,
    ) -> EvidenceWorkbook:
        self._validate_request(
            target_row_number=target_row_number,
            window_radius=window_radius,
            max_columns=max_columns,
            header_row_number=header_row_number,
        )
        try:
            object_stream = self.snapshot_store.open_object(content_sha256)
        except (OSError, ValueError) as exc:
            raise EvidencePreviewNotFoundError("snapshot content is unavailable") from exc

        with object_stream:
            try:
                with self._select_source(object_stream, member_name=member_name) as selected:
                    if selected.file_kind == "csv":
                        return self._preview_csv(
                            selected,
                            content_sha256=content_sha256,
                            sheet_name=sheet_name,
                            target_row_number=target_row_number,
                            window_radius=window_radius,
                            max_columns=max_columns,
                            target_column=target_column,
                            header_row_number=header_row_number,
                        )
                    return self._preview_xlsx(
                        selected,
                        content_sha256=content_sha256,
                        sheet_name=sheet_name,
                        target_row_number=target_row_number,
                        window_radius=window_radius,
                        max_columns=max_columns,
                        target_column=target_column,
                        header_row_number=header_row_number,
                    )
            except UnsafeZipError as exc:
                raise EvidencePreviewSecurityError("unsafe ZIP or XLSX package") from exc

    def _validate_request(
        self,
        *,
        target_row_number: int,
        window_radius: int,
        max_columns: int,
        header_row_number: int | None,
    ) -> None:
        if target_row_number <= 0:
            raise EvidencePreviewError("target_row_number must be a positive source row")
        if window_radius < 0 or window_radius > self.max_window_radius:
            raise EvidencePreviewError(
                f"window_radius must be between 0 and {self.max_window_radius}"
            )
        if max_columns <= 0 or max_columns > self.max_columns:
            raise EvidencePreviewError(f"max_columns must be between 1 and {self.max_columns}")
        if header_row_number is not None and header_row_number <= 0:
            raise EvidencePreviewError("header_row_number must be positive")
        if header_row_number is not None and target_row_number <= header_row_number:
            raise EvidencePreviewError("target row must be after the header row")

    @contextmanager
    def _select_source(
        self,
        object_stream: BinaryIO,
        *,
        member_name: str | None,
    ) -> Iterator[_SourceSelection]:
        object_stream.seek(0)
        prefix = object_stream.read(4)
        object_stream.seek(0)
        if not prefix.startswith(b"PK"):
            if member_name is not None:
                raise EvidencePreviewNotFoundError("member_name is only valid for ZIP snapshots")
            yield _SourceSelection(object_stream, "csv", None)
            return

        object_path = str(getattr(object_stream, "name", ""))
        if not object_path:
            raise EvidencePreviewSecurityError("snapshot ZIP is not seekable")
        policy = SafeZipPolicy(
            max_members=MAX_ARCHIVE_MEMBERS,
            max_total_uncompressed=MAX_ARCHIVE_TOTAL_BYTES,
            max_member_uncompressed=MAX_ARCHIVE_MEMBER_BYTES,
            max_compression_ratio=MAX_COMPRESSION_RATIO,
        )
        members = inspect_zip(object_path, policy=policy)
        member_names = {item.name.replace("\\", "/"): item for item in members}
        if self._is_xlsx_package(member_names):
            if member_name is not None:
                raise EvidencePreviewNotFoundError(
                    "member_name is not valid for a direct XLSX snapshot"
                )
            object_stream.seek(0)
            yield _SourceSelection(object_stream, "xlsx", None)
            return

        supported = [
            name
            for name in member_names
            if PurePosixPath(name).suffix.lower() in _SUPPORTED_MEMBER_SUFFIXES
        ]
        normalized_member = member_name.replace("\\", "/") if member_name is not None else None
        if normalized_member is None:
            if len(supported) != 1:
                raise EvidencePreviewError(
                    "ZIP snapshot requires an explicit member_name when it contains "
                    f"{len(supported)} supported tabular members"
                )
            normalized_member = supported[0]
        if normalized_member not in member_names:
            raise EvidencePreviewNotFoundError("requested ZIP member does not exist")
        suffix = PurePosixPath(normalized_member).suffix.lower()
        if suffix not in _SUPPORTED_MEMBER_SUFFIXES:
            raise EvidencePreviewError("requested ZIP member is not CSV or XLSX")

        raw_name = member_names[normalized_member].name
        with (
            zipfile.ZipFile(object_path, mode="r") as archive,
            archive.open(raw_name, mode="r") as source,
            tempfile.SpooledTemporaryFile(max_size=8 * 1024 * 1024, mode="w+b") as spool,
        ):
            copied = 0
            while chunk := source.read(_COPY_CHUNK_SIZE):
                copied += len(chunk)
                if copied > member_names[normalized_member].uncompressed_size:
                    raise EvidencePreviewSecurityError(
                        "ZIP member grew beyond its inspected size"
                    )
                spool.write(chunk)
            if copied != member_names[normalized_member].uncompressed_size:
                raise EvidencePreviewSecurityError(
                    "ZIP member size changed while reading"
                )
            spool.seek(0)
            yield _SourceSelection(
                cast(BinaryIO, spool),
                "csv" if suffix == ".csv" else "xlsx",
                normalized_member,
            )

    @staticmethod
    def _is_xlsx_package(member_names: Mapping[str, object]) -> bool:
        return "[Content_Types].xml" in member_names and "xl/workbook.xml" in member_names

    def _preview_csv(
        self,
        selected: _SourceSelection,
        *,
        content_sha256: str,
        sheet_name: str | None,
        target_row_number: int,
        window_radius: int,
        max_columns: int,
        target_column: str | int | None,
        header_row_number: int | None,
    ) -> EvidenceWorkbook:
        if sheet_name not in (None, "CSV"):
            raise EvidencePreviewNotFoundError("CSV snapshots only expose the CSV sheet")
        stream = selected.stream
        encoding, delimiter = self._csv_format(stream)
        stream.seek(0)
        text_stream = io.TextIOWrapper(stream, encoding=encoding, errors="strict", newline="")
        try:
            reader = csv.reader(text_stream, delimiter=delimiter)
            records = self._iter_csv_records(reader)
            header, candidates = self._find_csv_header(
                records,
                requested_header_row=header_row_number,
            )
            target, before, after = self._collect_csv_window(
                candidates,
                header=header,
                target_row_number=target_row_number,
                radius=window_radius,
            )
        except UnicodeDecodeError as exc:
            raise EvidencePreviewError("CSV contains invalid encoded text") from exc
        except csv.Error as exc:
            raise EvidencePreviewError("CSV structure is invalid") from exc
        finally:
            text_stream.detach()

        width = min(
            max(
                len(header.values),
                *(len(row.values) for row in before + [target] + after),
            ),
            max_columns,
        )
        columns = self._make_columns(header.values, width)
        target_column_index = self._resolve_target_column(target_column, columns)
        rows = tuple(
            EvidenceRow(
                source_row_number=record.source_row_number,
                source_end_row_number=record.source_end_row_number,
                data_row_number=index,
                cells=tuple(self._text_cell(value) for value in record.values[:width])
                + tuple(self._text_cell("") for _ in range(width - len(record.values))),
            )
            for index, record in (
                (
                    record.record_number - header.record_number,
                    record,
                )
                for record in before + [target] + after
            )
        )
        target_data_row_number = target.record_number - header.record_number
        window = EvidenceWindow(
            header_row_number=header.source_row_number,
            target_row_number=target_row_number,
            target_data_row_number=target_data_row_number,
            start_row_number=rows[0].source_row_number,
            end_row_number=rows[-1].source_end_row_number,
            columns=columns,
            rows=rows,
            target_column_index=target_column_index,
        )
        return EvidenceWorkbook(
            content_sha256=content_sha256,
            file_kind="csv",
            member_name=selected.member_name,
            sheet_names=("CSV",),
            sheet=EvidenceSheet(name="CSV", hidden=False, window=window),
        )

    @staticmethod
    def _csv_format(stream: BinaryIO) -> tuple[str, str]:
        stream.seek(0)
        first_line = stream.readline(64 * 1024)
        try:
            encoding = detect_csv_encoding(first_line).encoding
        except CsvEncodingError as exc:
            raise EvidencePreviewError("CSV encoding is unsupported") from exc
        stream.seek(0)
        sample = stream.read(128 * 1024)
        try:
            text = sample.decode(encoding, errors="ignore")
        except LookupError as exc:
            raise EvidencePreviewError("CSV encoding is unsupported") from exc
        try:
            delimiter = csv.Sniffer().sniff(text, delimiters=",\t;|").delimiter
        except csv.Error:
            counts = {candidate: text.count(candidate) for candidate in ",\t;|"}
            delimiter, count = max(counts.items(), key=lambda item: item[1])
            if count == 0:
                raise EvidencePreviewError("unable to detect CSV delimiter") from None
        return encoding, str(delimiter)

    @staticmethod
    def _iter_csv_records(reader: _CsvReader) -> Iterator[_CsvRecord]:
        previous_end = 0
        for record_number, values in enumerate(reader, start=1):
            current_end = reader.line_num
            yield _CsvRecord(
                record_number=record_number,
                source_row_number=previous_end + 1,
                source_end_row_number=current_end,
                values=tuple(values),
            )
            previous_end = current_end

    @staticmethod
    def _find_csv_header(
        records: Iterator[_CsvRecord],
        *,
        requested_header_row: int | None,
    ) -> tuple[_CsvRecord, Iterator[_CsvRecord]]:
        buffered: list[_CsvRecord] = []
        header: _CsvRecord | None = None
        for record in records:
            if record.source_row_number > MAX_HEADER_ROWS and requested_header_row is None:
                break
            buffered.append(record)
            if requested_header_row is not None:
                if record.source_row_number <= requested_header_row <= record.source_end_row_number:
                    header = record
                    break
            elif len([value for value in record.values if value.strip()]) >= 2:
                header = record
                break
        if header is None:
            raise EvidencePreviewNotFoundError("no usable CSV header row was found")

        def remaining() -> Iterator[_CsvRecord]:
            for record in buffered:
                if record.source_row_number > header.source_end_row_number:
                    yield record
            yield from records

        return header, remaining()

    @staticmethod
    def _collect_csv_window(
        records: Iterator[_CsvRecord],
        *,
        header: _CsvRecord,
        target_row_number: int,
        radius: int,
    ) -> tuple[_CsvRecord, list[_CsvRecord], list[_CsvRecord]]:
        if target_row_number <= header.source_end_row_number:
            raise EvidencePreviewError("target row must be after the header row")
        before: deque[_CsvRecord] = deque(maxlen=radius)
        target: _CsvRecord | None = None
        after: list[_CsvRecord] = []
        for record in records:
            if target is None:
                if record.source_row_number <= target_row_number <= record.source_end_row_number:
                    target = record
                    continue
                if record.source_row_number > target_row_number:
                    break
                before.append(record)
                continue
            if len(after) >= radius:
                break
            after.append(record)
        if target is None:
            raise EvidencePreviewNotFoundError("target CSV source row does not exist")
        return target, list(before), after

    def _preview_xlsx(
        self,
        selected: _SourceSelection,
        *,
        content_sha256: str,
        sheet_name: str | None,
        target_row_number: int,
        window_radius: int,
        max_columns: int,
        target_column: str | int | None,
        header_row_number: int | None,
    ) -> EvidenceWorkbook:
        selected.stream.seek(0)
        try:
            workbook = load_workbook(
                selected.stream,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise EvidencePreviewError("invalid or unsupported XLSX workbook") from exc
        try:
            sheet_names = tuple(workbook.sheetnames)
            if not sheet_name:
                raise EvidencePreviewNotFoundError(
                    "source XLSX worksheet was not frozen"
                )
            selected_sheet_name = sheet_name
            if selected_sheet_name is None or selected_sheet_name not in workbook:
                raise EvidencePreviewNotFoundError("requested XLSX sheet does not exist")
            worksheet = workbook[selected_sheet_name]
            header_number, header_values = self._find_xlsx_header(
                worksheet,
                requested_header_row=header_row_number,
                max_columns=max_columns,
            )
            if target_row_number <= header_number:
                raise EvidencePreviewError("target row must be after the header row")
            start = max(header_number + 1, target_row_number - window_radius)
            end = target_row_number + window_radius
            raw_rows = list(
                worksheet.iter_rows(
                    min_row=start,
                    max_row=end,
                    max_col=max_columns,
                    values_only=False,
                )
            )
            target_offset = target_row_number - start
            if target_offset >= len(raw_rows) or self._xlsx_row_is_absent(raw_rows[target_offset]):
                raise EvidencePreviewNotFoundError("target XLSX row does not exist")
            while raw_rows and self._xlsx_row_is_absent(raw_rows[-1]):
                raw_rows.pop()
            width = min(
                max(
                    len(header_values),
                    max(
                        (
                            self._last_nonempty_xlsx_column(row)
                            for row in raw_rows
                        ),
                        default=0,
                    ),
                ),
                max_columns,
            )
            columns = self._make_columns(header_values, width)
            target_column_index = self._resolve_target_column(target_column, columns)
            rows = tuple(
                EvidenceRow(
                    source_row_number=start + offset,
                    source_end_row_number=start + offset,
                    data_row_number=start + offset - header_number,
                    cells=tuple(self._xlsx_cell(cell) for cell in row[:width]),
                )
                for offset, row in enumerate(raw_rows)
            )
            window = EvidenceWindow(
                header_row_number=header_number,
                target_row_number=target_row_number,
                target_data_row_number=target_row_number - header_number,
                start_row_number=rows[0].source_row_number,
                end_row_number=rows[-1].source_end_row_number,
                columns=columns,
                rows=rows,
                target_column_index=target_column_index,
            )
            return EvidenceWorkbook(
                content_sha256=content_sha256,
                file_kind="xlsx",
                member_name=selected.member_name,
                sheet_names=sheet_names,
                sheet=EvidenceSheet(
                    name=worksheet.title,
                    hidden=worksheet.sheet_state != "visible",
                    window=window,
                ),
            )
        finally:
            workbook.close()

    @staticmethod
    def _find_xlsx_header(
        worksheet: Any,
        *,
        requested_header_row: int | None,
        max_columns: int,
    ) -> tuple[int, tuple[str, ...]]:
        requested = (
            range(requested_header_row, requested_header_row + 1)
            if requested_header_row is not None
            else range(1, MAX_HEADER_ROWS + 1)
        )
        for row_number in requested:
            rows = list(
                worksheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                    max_col=max_columns,
                    values_only=True,
                )
            )
            if not rows:
                continue
            values = tuple("" if value is None else str(value).strip() for value in rows[0])
            while values and not values[-1]:
                values = values[:-1]
            if requested_header_row is not None or len([value for value in values if value]) >= 2:
                return row_number, values
        raise EvidencePreviewNotFoundError("no usable XLSX header row was found")

    @staticmethod
    def _make_columns(
        header_values: tuple[str, ...],
        width: int,
    ) -> tuple[EvidenceColumn, ...]:
        return tuple(
            EvidenceColumn(
                index=index,
                source_label=header_values[index] if index < len(header_values) else "",
                label=(
                    header_values[index]
                    if index < len(header_values) and header_values[index]
                    else f"列 {index + 1}"
                ),
            )
            for index in range(width)
        )

    @staticmethod
    def _resolve_target_column(
        target_column: str | int | None,
        columns: tuple[EvidenceColumn, ...],
    ) -> int | None:
        if target_column is None:
            return None
        if isinstance(target_column, int):
            if target_column < 0 or target_column >= len(columns):
                raise EvidencePreviewNotFoundError("target column index does not exist")
            return target_column
        normalized = target_column.strip().casefold()
        matches = [
            column.index
            for column in columns
            if column.source_label.strip().casefold() == normalized
        ]
        if len(matches) != 1:
            raise EvidencePreviewNotFoundError(
                "target column name is missing or not unique"
            )
        return matches[0]

    @staticmethod
    def _text_cell(value: str) -> EvidenceCell:
        if value == "":
            return EvidenceCell(value=None, value_kind="blank")
        return EvidenceCell(value=value, value_kind="text")

    @staticmethod
    def _xlsx_cell(cell: Cell) -> EvidenceCell:
        value = cell.value
        if cell.data_type == "f":
            formula = str(value) if value is not None else ""
            return EvidenceCell(
                value=None,
                value_kind="formula",
                formula=formula,
                deterministic=False,
            )
        if value is None:
            return EvidenceCell(value=None, value_kind="blank")
        if cell.data_type == "e":
            return EvidenceCell(value=str(value), value_kind="error", deterministic=False)
        if isinstance(value, bool):
            return EvidenceCell(value="true" if value else "false", value_kind="boolean")
        if isinstance(value, datetime):
            return EvidenceCell(value=value.isoformat(), value_kind="datetime")
        if isinstance(value, date):
            return EvidenceCell(value=value.isoformat(), value_kind="date")
        if isinstance(value, time):
            return EvidenceCell(value=value.isoformat(), value_kind="time")
        if isinstance(value, (int, float, Decimal)):
            return EvidenceCell(value=str(value), value_kind="number")
        return EvidenceCell(value=str(value), value_kind="text")

    @staticmethod
    def _xlsx_row_is_absent(row: tuple[Cell, ...]) -> bool:
        return all(cell.value is None for cell in row)

    @staticmethod
    def _last_nonempty_xlsx_column(row: tuple[Cell, ...]) -> int:
        for index in range(len(row), 0, -1):
            if row[index - 1].value is not None:
                return index
        return 0
