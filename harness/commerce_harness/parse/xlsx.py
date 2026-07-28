from __future__ import annotations

import io
import os
from pathlib import Path

from openpyxl import load_workbook

from .models import HeaderCandidate, TabularProfile
from .zip_safe import SafeZipPolicy, UnsafeZipError, inspect_zip


class XlsxStructureError(ValueError):
    pass


def profile_xlsx(
    source: bytes | str | os.PathLike[str],
    *,
    member_name: str | None = None,
    max_header_rows: int = 20,
    max_columns: int = 256,
) -> TabularProfile:
    workbook_source = io.BytesIO(source) if isinstance(source, bytes) else Path(source)
    try:
        inspect_zip(
            source,
            policy=SafeZipPolicy(
                max_members=2048,
                max_total_uncompressed=1024 * 1024 * 1024,
                max_member_uncompressed=512 * 1024 * 1024,
                max_compression_ratio=250.0,
            ),
        )
    except (UnsafeZipError, OSError) as exc:
        raise XlsxStructureError("unsafe XLSX package") from exc
    try:
        workbook = load_workbook(
            workbook_source,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as exc:
        raise XlsxStructureError("invalid or unsupported XLSX workbook") from exc

    candidates: list[HeaderCandidate] = []
    sheet_names = tuple(workbook.sheetnames)
    try:
        for worksheet in workbook.worksheets:
            if hasattr(worksheet, "reset_dimensions"):
                worksheet.reset_dimensions()
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=max_header_rows,
                    max_col=max_columns,
                    values_only=True,
                ),
                start=1,
            ):
                headers = tuple("" if value is None else str(value).strip() for value in row)
                while headers and not headers[-1]:
                    headers = headers[:-1]
                nonempty = [value for value in headers if value]
                if len(nonempty) >= 2 and len(set(nonempty)) == len(nonempty):
                    candidates.append(
                        HeaderCandidate(
                            headers=headers,
                            header_row=row_number,
                            sheet_name=worksheet.title,
                            sheet_hidden=worksheet.sheet_state != "visible",
                            member_name=member_name,
                        )
                    )
    finally:
        workbook.close()

    if not candidates:
        raise XlsxStructureError("no plausible header row found in XLSX")
    return TabularProfile(
        file_kind="xlsx",
        candidates=tuple(candidates),
        sheet_names=sheet_names,
    )
