from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path, PureWindowsPath
from typing import Any

import pyarrow as pa
from openpyxl import load_workbook

from .memory.database import DuckDBMemory
from .workbench import WorkbenchPaths

MONEY_QUANTUM = Decimal("0.0001")
FORMULA_TOLERANCE = Decimal("0.0200")
ASSIGNMENT_START_PERIOD = "2602"

REFERENCE_HEADERS = {
    "person": "姓名",
    "product": "宝贝编码",
    "collected": "交易收款",
    "refund": "交易退款",
    "compensation": "交易赔付",
    "software_fee": "软件服务费",
    "marketing_fee": "营销费用",
    "shipping_fee": "发货运费",
    "product_cost": "订单成本",
    "reship_cost": "补发成本",
    "principal_commission": "本金佣金",
    "procurement": "代购代发",
    "gross_profit": "店铺毛利",
    "advertising_fee": "广告费",
    "store_profit": "店铺利润",
}

REFERENCE_REQUIRED_KEYS = frozenset(
    {
        "product",
        "collected",
        "refund",
        "gross_profit",
        "advertising_fee",
        "store_profit",
    }
)


@dataclass(frozen=True, slots=True)
class PerformanceSyncResult:
    imported_snapshots: int
    skipped_snapshots: int
    employee_rows: int
    assignment_rows: int
    reference_rows: int
    issue_count: int


def _stable_id(prefix: str, *parts: str) -> str:
    payload = "\x1f".join(parts).encode("utf-8")
    return f"{prefix}_{hashlib.sha256(payload).hexdigest()[:32]}"


def _normalized_text(value: object) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).split()).strip()


def _normalized_alias(value: object) -> str:
    return _normalized_text(value).casefold()


def _money(value: object) -> Decimal:
    text = _normalized_text(value).replace(",", "")
    if not text:
        return Decimal("0.0000")
    try:
        return Decimal(text).quantize(MONEY_QUANTUM, rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError(f"金额格式无法识别: {text[:40]}") from exc


def calculate_reference_formula(row: dict[str, Decimal]) -> tuple[Decimal, Decimal]:
    gross = sum(
        (
            row["collected"],
            row["refund"],
            row["compensation"],
            row["software_fee"],
            row["marketing_fee"],
            row["shipping_fee"],
            row["product_cost"],
            row["reship_cost"],
            row["principal_commission"],
            row["procurement"],
        ),
        Decimal("0.0000"),
    ).quantize(MONEY_QUANTUM)
    gross_residual = (gross - row["gross_profit"]).quantize(MONEY_QUANTUM)
    profit_residual = (
        row["gross_profit"] + row["advertising_fee"] - row["store_profit"]
    ).quantize(MONEY_QUANTUM)
    return gross_residual, profit_residual


def _latest_relevant_snapshots(database: DuckDBMemory) -> list[dict[str, str]]:
    rows = database.execute(
        """
        SELECT snapshot_id, object_uri, source_uri, original_name,
               cast(captured_at AS VARCHAR)
        FROM source_snapshot
        ORDER BY captured_at DESC, snapshot_id DESC
        """
    ).fetchall()
    latest: dict[str, dict[str, str]] = {}
    for snapshot_id, object_uri, source_uri, original_name, captured_at in rows:
        uri = str(source_uri or "")
        normalized = uri.replace("/", "\\").casefold()
        source_kind = ""
        if normalized.endswith("\\工资\\员工信息表.xlsx"):
            source_kind = "employee_master"
        elif normalized.endswith("\\其他\\运营链接.xlsx"):
            source_kind = "operator_assignment"
        elif (
            "\\工资\\2026\\阿里单算\\" in normalized
            or "\\工资\\2026\\阿里合算\\" in normalized
        ) and normalized.endswith(".csv"):
            source_kind = "performance_reference"
        if not source_kind or uri in latest:
            continue
        latest[uri] = {
            "snapshot_id": str(snapshot_id),
            "object_uri": str(object_uri),
            "source_uri": uri,
            "original_name": str(original_name or _source_name(uri)),
            "captured_at": str(captured_at),
            "source_kind": source_kind,
        }
    return sorted(
        latest.values(),
        key=lambda item: (item["source_kind"], item["source_uri"].casefold()),
    )


def _source_name(source_uri: str) -> str:
    value = source_uri.split("://", 1)[-1]
    return PureWindowsPath(value).name


def _ensure_person(
    connection: Any,
    *,
    enterprise_id: str,
    name: object,
    status: str,
    snapshot_id: str,
    row_no: int,
    department: object = None,
    employment_type: object = None,
    cache: dict[str, str] | None = None,
) -> str | None:
    display_name = _normalized_text(name)
    normalized = _normalized_alias(display_name)
    if not normalized:
        return None
    if cache is not None and normalized in cache:
        return cache[normalized]
    person_id = _stable_id("person", enterprise_id, normalized)
    checksum = hashlib.sha256(f"{enterprise_id}\x1f{normalized}".encode()).hexdigest()
    existing = connection.execute(
        "SELECT person_id FROM person_identity WHERE person_id = ?",
        [person_id],
    ).fetchone()
    if existing is None:
        connection.execute(
            """
            INSERT INTO person_identity (
                person_id, enterprise_id, display_name, department, employment_type,
                status, source_snapshot_id, source_row_no, identity_checksum
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                person_id,
                enterprise_id,
                display_name,
                _normalized_text(department) or None,
                _normalized_text(employment_type) or None,
                status,
                snapshot_id,
                row_no,
                checksum,
            ],
        )
    alias_id = _stable_id("alias", enterprise_id, normalized)
    connection.execute(
        """
        INSERT INTO person_alias (
            alias_id, enterprise_id, person_id, alias_text, normalized_alias,
            effective_from, source_snapshot_id, source_row_no, status
        )
        VALUES (?, ?, ?, ?, ?, DATE '2024-01-01', ?, ?, 'active')
        ON CONFLICT (alias_id) DO NOTHING
        """,
        [
            alias_id,
            enterprise_id,
            person_id,
            display_name,
            normalized,
            snapshot_id,
            row_no,
        ],
    )
    if cache is not None:
        cache[normalized] = person_id
    return person_id


def _ensure_product(
    connection: Any,
    enterprise_id: str,
    value: object,
    *,
    cache: dict[str, str] | None = None,
) -> str | None:
    code = _normalized_text(value)
    if not code:
        return None
    normalized = code.casefold()
    if cache is not None and normalized in cache:
        return cache[normalized]
    product_id = _stable_id("product", enterprise_id, normalized)
    existing = connection.execute(
        "SELECT product_id FROM canonical_product WHERE product_id = ?",
        [product_id],
    ).fetchone()
    if existing is None:
        connection.execute(
            """
            INSERT INTO canonical_product (
                product_id, enterprise_id, merchant_product_code, status
            )
            VALUES (?, ?, ?, 'provisional')
            """,
            [product_id, enterprise_id, code],
        )
    if cache is not None:
        cache[normalized] = product_id
    return product_id


def _bulk_insert(
    connection: Any,
    *,
    table: str,
    columns: list[str],
    rows: list[list[object]],
    conflict_columns: list[str],
) -> None:
    if not rows:
        return
    relation_name = f"_performance_stage_{table}"
    payload = pa.Table.from_pylist(
        [dict(zip(columns, row, strict=True)) for row in rows]
    )
    connection.register(relation_name, payload)
    try:
        column_sql = ", ".join(columns)
        conflict_sql = ", ".join(conflict_columns)
        connection.execute(
            f"""
            INSERT INTO {table} ({column_sql})
            SELECT {column_sql} FROM {relation_name}
            ON CONFLICT ({conflict_sql}) DO NOTHING
            """
        )
    finally:
        connection.unregister(relation_name)


def _bulk_ensure_products(
    connection: Any,
    *,
    enterprise_id: str,
    values: list[object],
    cache: dict[str, str],
) -> None:
    rows: list[list[object]] = []
    for value in values:
        code = _normalized_text(value)
        normalized = code.casefold()
        if not normalized or normalized in cache:
            continue
        product_id = _stable_id("product", enterprise_id, normalized)
        cache[normalized] = product_id
        rows.append([product_id, enterprise_id, code, "provisional"])
    _bulk_insert(
        connection,
        table="canonical_product",
        columns=[
            "product_id",
            "enterprise_id",
            "merchant_product_code",
            "status",
        ],
        rows=rows,
        conflict_columns=["product_id"],
    )


def _period_dates(period_token: str) -> tuple[date, date]:
    if not re.fullmatch(r"\d{4}", period_token):
        raise ValueError(f"无效月份: {period_token}")
    year = 2000 + int(period_token[:2])
    month = int(period_token[2:])
    return date(year, month, 1), date(year, month, monthrange(year, month)[1])


def _csv_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                reader = csv.reader(stream)
                rows = list(reader)
            if not rows:
                raise ValueError("CSV 为空")
            return rows[0], rows[1:]
        except UnicodeError:
            continue
    raise ValueError("CSV 编码无法识别")


def _record_import(
    database: DuckDBMemory,
    *,
    enterprise_id: str,
    snapshot: dict[str, str],
    loader: Any,
) -> tuple[int, int, bool, bool]:
    already = database.execute(
        """
        SELECT status
        FROM performance_source_import
        WHERE enterprise_id = ? AND snapshot_id = ? AND source_kind = ?
        """,
        [enterprise_id, snapshot["snapshot_id"], snapshot["source_kind"]],
    ).fetchone()
    if already is not None and str(already[0]) == "succeeded":
        return 0, 0, True, True
    import_id = _stable_id(
        "performance_import",
        enterprise_id,
        snapshot["snapshot_id"],
        snapshot["source_kind"],
    )
    database.execute(
        """
        INSERT INTO performance_source_import (
            import_id, enterprise_id, snapshot_id, source_kind, status
        )
        VALUES (?, ?, ?, ?, 'running')
        ON CONFLICT (import_id) DO UPDATE SET
            status = 'running', error_detail = NULL, started_at = now(),
            finished_at = NULL
        """,
        [
            import_id,
            enterprise_id,
            snapshot["snapshot_id"],
            snapshot["source_kind"],
        ],
    )
    try:
        row_count, issue_count = loader(snapshot)
    except Exception as exc:
        database.execute(
            """
            UPDATE performance_source_import
            SET status = 'failed', finished_at = current_timestamp, error_detail = ?
            WHERE import_id = ?
            """,
            [str(exc)[:2000], import_id],
        )
        return 0, 1, False, False
    database.execute(
        """
        UPDATE performance_source_import
        SET status = 'succeeded', row_count = ?, issue_count = ?,
            finished_at = current_timestamp,
            metrics_json = ?
        WHERE import_id = ?
        """,
        [
            row_count,
            issue_count,
            json.dumps({"source_uri_sha256": hashlib.sha256(
                snapshot["source_uri"].encode("utf-8")
            ).hexdigest()}),
            import_id,
        ],
    )
    return row_count, issue_count, False, True


def sync_performance_sources(
    workbench: WorkbenchPaths,
    *,
    enterprise_id: str,
) -> PerformanceSyncResult:
    employee_rows = 0
    assignment_rows = 0
    reference_rows = 0
    issues = 0
    imported = 0
    skipped = 0
    with DuckDBMemory(workbench.database) as database:
        database.initialize()
        # The agreed product scope starts in 2026-02. January files remain in
        # the immutable snapshot store as evidence, but are not exposed as
        # performance facts or active responsibility assignments.
        database.execute(
            """
            DELETE FROM performance_reference_fact
            WHERE enterprise_id = ? AND period_token < ?
            """,
            [enterprise_id, ASSIGNMENT_START_PERIOD],
        )
        database.execute(
            """
            DELETE FROM responsibility_assignment_version
            WHERE enterprise_id = ?
              AND effective_to < DATE '2026-02-01'
            """,
            [enterprise_id],
        )
        snapshots = _latest_relevant_snapshots(database)
        person_cache = {
            str(normalized_alias): str(person_id)
            for normalized_alias, person_id in database.execute(
                """
                SELECT normalized_alias, person_id
                FROM person_alias
                WHERE enterprise_id = ? AND status = 'active'
                """,
                [enterprise_id],
            ).fetchall()
        }
        product_cache = {
            _normalized_alias(code): str(product_id)
            for code, product_id in database.execute(
                """
                SELECT merchant_product_code, product_id
                FROM canonical_product
                WHERE enterprise_id = ?
                """,
                [enterprise_id],
            ).fetchall()
        }

        def import_employee(snapshot: dict[str, str]) -> tuple[int, int]:
            path = Path(snapshot["object_uri"])
            workbook = load_workbook(
                io.BytesIO(path.read_bytes()),
                read_only=True,
                data_only=True,
            )
            worksheet = workbook.worksheets[0]
            rows = 0
            issue_count = 0
            with database.transaction() as connection:
                for row_no, values in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    if not values or not _normalized_text(values[0]):
                        continue
                    person_id = _ensure_person(
                        connection,
                        enterprise_id=enterprise_id,
                        name=values[0],
                        status="active",
                        snapshot_id=snapshot["snapshot_id"],
                        row_no=row_no,
                        department=values[2] if len(values) > 2 else None,
                        employment_type=values[3] if len(values) > 3 else None,
                        cache=person_cache,
                    )
                    if person_id is None:
                        issue_count += 1
                    else:
                        rows += 1
            return rows, issue_count

        def import_operator(snapshot: dict[str, str]) -> tuple[int, int]:
            path = Path(snapshot["object_uri"])
            workbook = load_workbook(
                io.BytesIO(path.read_bytes()),
                read_only=True,
                data_only=True,
            )
            worksheet = workbook.worksheets[0]
            iterator = worksheet.iter_rows(values_only=True)
            headers = next(iterator)
            month_columns = {
                str(value)[:4]: index
                for index, value in enumerate(headers)
                if isinstance(value, str)
                and re.fullmatch(r"26(0[1-9]|1[0-2])姓名", value)
                and str(value)[:4] >= ASSIGNMENT_START_PERIOD
            }
            candidates: dict[
                tuple[str, str], list[tuple[str, int]]
            ] = {}
            for row_no, values in enumerate(iterator, start=2):
                if not values:
                    continue
                product_code = _normalized_text(values[0])
                if not product_code:
                    continue
                for period_token, index in month_columns.items():
                    name = _normalized_text(values[index] if index < len(values) else None)
                    if name:
                        candidates.setdefault((period_token, product_code), []).append(
                            (name, row_no)
                        )
            inserted = 0
            issue_count = 0
            source_key = _stable_id("operator_source", snapshot["source_uri"])
            with database.transaction() as connection:
                connection.execute(
                    """
                    UPDATE responsibility_assignment_version
                    SET status = 'superseded'
                    WHERE source_kind = ? AND source_snapshot_id <> ?
                      AND status = 'active'
                    """,
                    [source_key, snapshot["snapshot_id"]],
                )
                unique_people: dict[str, tuple[str, int]] = {}
                for assignments in candidates.values():
                    for name, row_no in assignments:
                        unique_people.setdefault(
                            _normalized_alias(name),
                            (name, row_no),
                        )
                for name, row_no in unique_people.values():
                    _ensure_person(
                        connection,
                        enterprise_id=enterprise_id,
                        name=name,
                        status="provisional",
                        snapshot_id=snapshot["snapshot_id"],
                        row_no=row_no,
                        cache=person_cache,
                    )
                _bulk_ensure_products(
                    connection,
                    enterprise_id=enterprise_id,
                    values=[product_code for _, product_code in candidates],
                    cache=product_cache,
                )
                assignment_parameters: list[list[object]] = []
                for (period_token, product_code), assignments in candidates.items():
                    people = {_normalized_alias(name) for name, _ in assignments}
                    conflict = len(people) > 1
                    name, row_no = assignments[0]
                    person_id = _ensure_person(
                        connection,
                        enterprise_id=enterprise_id,
                        name=name,
                        status="provisional",
                        snapshot_id=snapshot["snapshot_id"],
                        row_no=row_no,
                        cache=person_cache,
                    )
                    product_id = _ensure_product(
                        connection,
                        enterprise_id,
                        product_code,
                        cache=product_cache,
                    )
                    if person_id is None or product_id is None:
                        issue_count += 1
                        continue
                    effective_from, effective_to = _period_dates(period_token)
                    checksum = hashlib.sha256(
                        "\x1f".join(
                            (
                                enterprise_id,
                                person_id,
                                product_id,
                                period_token,
                                source_key,
                            )
                        ).encode("utf-8")
                    ).hexdigest()
                    assignment_id = _stable_id(
                        "assignment",
                        snapshot["snapshot_id"],
                        period_token,
                        product_id,
                        person_id,
                    )
                    assignment_parameters.append(
                        [
                            assignment_id,
                            enterprise_id,
                            person_id,
                            product_id,
                            effective_from,
                            effective_to,
                            1,
                            "conflict" if conflict else "active",
                            source_key,
                            snapshot["snapshot_id"],
                            worksheet.title,
                            row_no,
                            checksum,
                        ]
                    )
                    inserted += 1
                    if conflict:
                        issue_count += 1
                if assignment_parameters:
                    _bulk_insert(
                        connection,
                        table="responsibility_assignment_version",
                        columns=[
                            "assignment_id",
                            "enterprise_id",
                            "person_id",
                            "product_id",
                            "effective_from",
                            "effective_to",
                            "version",
                            "status",
                            "source_kind",
                            "source_snapshot_id",
                            "source_sheet",
                            "source_row_no",
                            "checksum_sha256",
                        ],
                        rows=assignment_parameters,
                        conflict_columns=["assignment_id"],
                    )
            return inserted, issue_count

        def import_reference(snapshot: dict[str, str]) -> tuple[int, int]:
            path = Path(snapshot["object_uri"])
            headers, rows = _csv_rows(path)
            if len(headers) < 2:
                raise ValueError("绩效对照文件缺少表头")
            person_header = headers[0]
            match = re.match(r"(?P<period>\d{4})姓名$", person_header)
            if match is None:
                # The same folders also contain store-level monthly summaries.
                # They are valid historical assets, but they have neither a
                # person-product grain nor a safe deterministic attribution.
                # Do not report them as broken personnel performance files.
                return 0, 0
            period_token = match.group("period")
            if period_token < ASSIGNMENT_START_PERIOD:
                return 0, 0
            required = {
                person_header,
                *(
                    header
                    for key, header in REFERENCE_HEADERS.items()
                    if key in REFERENCE_REQUIRED_KEYS
                ),
            }
            missing = sorted(required - set(headers))
            if missing:
                raise ValueError(f"绩效对照文件缺少字段: {', '.join(missing)}")
            index = {value: headers.index(value) for value in headers}
            source_uri = snapshot["source_uri"].replace("/", "\\")
            mode = "combined" if "\\阿里合算\\" in source_uri else "single"
            store_name = Path(snapshot["original_name"]).stem
            prepared: list[dict[str, Any]] = []
            assignments: dict[tuple[str, str], set[str]] = {}
            for row_no, values in enumerate(rows, start=2):
                def cell(header: str, row_values: list[str] = values) -> str:
                    position = index.get(header)
                    if position is None:
                        return ""
                    return (
                        row_values[position] if position < len(row_values) else ""
                    )

                person_name = _normalized_text(cell(person_header))
                product_code = _normalized_text(cell(REFERENCE_HEADERS["product"]))
                if not person_name or not product_code:
                    continue
                amounts = {
                    key: _money(cell(header))
                    for key, header in REFERENCE_HEADERS.items()
                    if key not in {"person", "product"}
                }
                gross_residual, profit_residual = calculate_reference_formula(amounts)
                prepared.append(
                    {
                        "row_no": row_no,
                        "person_name": person_name,
                        "product_code": product_code,
                        "amounts": amounts,
                        "gross_residual": gross_residual,
                        "profit_residual": profit_residual,
                    }
                )
                assignments.setdefault((store_name, product_code), set()).add(
                    _normalized_alias(person_name)
                )
            inserted = 0
            issue_count = 0
            effective_from, effective_to = _period_dates(period_token)
            source_key = _stable_id("reference_source", snapshot["source_uri"])
            with database.transaction() as connection:
                connection.execute(
                    """
                    UPDATE responsibility_assignment_version
                    SET status = 'superseded'
                    WHERE source_kind = ? AND source_snapshot_id <> ?
                      AND status = 'active'
                    """,
                    [source_key, snapshot["snapshot_id"]],
                )
                unique_people: dict[str, tuple[str, int]] = {}
                for item in prepared:
                    unique_people.setdefault(
                        _normalized_alias(item["person_name"]),
                        (item["person_name"], item["row_no"]),
                    )
                for name, row_no in unique_people.values():
                    _ensure_person(
                        connection,
                        enterprise_id=enterprise_id,
                        name=name,
                        status="provisional",
                        snapshot_id=snapshot["snapshot_id"],
                        row_no=row_no,
                        cache=person_cache,
                    )
                _bulk_ensure_products(
                    connection,
                    enterprise_id=enterprise_id,
                    values=[item["product_code"] for item in prepared],
                    cache=product_cache,
                )
                assignment_parameters: list[list[object]] = []
                fact_parameters: list[list[object]] = []
                for item in prepared:
                    person_id = _ensure_person(
                        connection,
                        enterprise_id=enterprise_id,
                        name=item["person_name"],
                        status="provisional",
                        snapshot_id=snapshot["snapshot_id"],
                        row_no=item["row_no"],
                        cache=person_cache,
                    )
                    product_id = _ensure_product(
                        connection,
                        enterprise_id,
                        item["product_code"],
                        cache=product_cache,
                    )
                    if person_id is None or product_id is None:
                        issue_count += 1
                        continue
                    conflict = (
                        len(assignments[(store_name, item["product_code"])]) > 1
                    )
                    assignment_checksum = hashlib.sha256(
                        "\x1f".join(
                            (
                                enterprise_id,
                                person_id,
                                product_id,
                                store_name,
                                period_token,
                                source_key,
                            )
                        ).encode("utf-8")
                    ).hexdigest()
                    assignment_parameters.append(
                        [
                            _stable_id(
                                "assignment",
                                snapshot["snapshot_id"],
                                store_name,
                                period_token,
                                product_id,
                                person_id,
                            ),
                            enterprise_id,
                            person_id,
                            product_id,
                            store_name,
                            effective_from,
                            effective_to,
                            1,
                            "conflict" if conflict else "active",
                            source_key,
                            snapshot["snapshot_id"],
                            item["row_no"],
                            assignment_checksum,
                        ]
                    )
                    amounts = item["amounts"]
                    passed = (
                        abs(item["gross_residual"]) <= FORMULA_TOLERANCE
                        and item["profit_residual"] == Decimal("0.0000")
                    )
                    fact_parameters.append(
                        [
                            _stable_id(
                                "reference",
                                snapshot["snapshot_id"],
                                str(item["row_no"]),
                            ),
                            enterprise_id,
                            period_token,
                            store_name,
                            person_id,
                            product_id,
                            mode,
                            amounts["collected"],
                            amounts["refund"],
                            amounts["compensation"],
                            amounts["software_fee"],
                            amounts["marketing_fee"],
                            amounts["shipping_fee"],
                            amounts["product_cost"],
                            amounts["reship_cost"],
                            amounts["principal_commission"],
                            amounts["procurement"],
                            amounts["gross_profit"],
                            amounts["advertising_fee"],
                            amounts["store_profit"],
                            item["gross_residual"],
                            item["profit_residual"],
                            "passed" if passed else "failed",
                            snapshot["snapshot_id"],
                            item["row_no"],
                        ]
                    )
                    inserted += 1
                    if conflict or not passed:
                        issue_count += 1
                if assignment_parameters:
                    _bulk_insert(
                        connection,
                        table="responsibility_assignment_version",
                        columns=[
                            "assignment_id",
                            "enterprise_id",
                            "person_id",
                            "product_id",
                            "store_name",
                            "effective_from",
                            "effective_to",
                            "version",
                            "status",
                            "source_kind",
                            "source_snapshot_id",
                            "source_row_no",
                            "checksum_sha256",
                        ],
                        rows=assignment_parameters,
                        conflict_columns=["assignment_id"],
                    )
                if fact_parameters:
                    _bulk_insert(
                        connection,
                        table="performance_reference_fact",
                        columns=[
                            "reference_id",
                            "enterprise_id",
                            "period_token",
                            "store_name",
                            "person_id",
                            "product_id",
                            "calculation_mode",
                            "collected_amount",
                            "refund_amount",
                            "compensation_amount",
                            "software_fee",
                            "marketing_fee",
                            "shipping_fee",
                            "product_cost",
                            "reship_cost",
                            "principal_commission",
                            "procurement_amount",
                            "gross_profit",
                            "advertising_fee",
                            "store_profit",
                            "gross_formula_residual",
                            "profit_formula_residual",
                            "validation_status",
                            "source_snapshot_id",
                            "source_row_no",
                        ],
                        rows=fact_parameters,
                        conflict_columns=["source_snapshot_id", "source_row_no"],
                    )
            return inserted, issue_count

        for snapshot in snapshots:
            if snapshot["source_kind"] == "employee_master":
                loader = import_employee
            elif snapshot["source_kind"] == "operator_assignment":
                loader = import_operator
            else:
                loader = import_reference
            rows, item_issues, was_skipped, succeeded = _record_import(
                database,
                enterprise_id=enterprise_id,
                snapshot=snapshot,
                loader=loader,
            )
            normalized_uri = snapshot["source_uri"].replace("/", "\\")
            if (
                snapshot["source_kind"] == "performance_reference"
                and "\\2026\\阿里" in normalized_uri
                and "\\1月\\" in normalized_uri
            ):
                database.execute(
                    """
                    UPDATE performance_source_import
                    SET row_count = 0, issue_count = 0,
                        metrics_json = json_merge_patch(
                            coalesce(metrics_json, '{}'),
                            '{"excluded_before_activation": true}'
                        )
                    WHERE enterprise_id = ? AND snapshot_id = ?
                      AND source_kind = 'performance_reference'
                    """,
                    [enterprise_id, snapshot["snapshot_id"]],
                )
            skipped += was_skipped
            if was_skipped:
                continue
            issues += item_issues
            if not succeeded:
                continue
            imported += 1
            if snapshot["source_kind"] == "employee_master":
                employee_rows += rows
            elif snapshot["source_kind"] == "operator_assignment":
                assignment_rows += rows
            else:
                reference_rows += rows

    return PerformanceSyncResult(
        imported_snapshots=imported,
        skipped_snapshots=skipped,
        employee_rows=employee_rows,
        assignment_rows=assignment_rows,
        reference_rows=reference_rows,
        issue_count=issues,
    )
