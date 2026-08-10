"""原语二：解析。把文件变成带行号的原始行。

行号是后续全部证据链的锚点，所以解析阶段一个字节都不能猜。

这里固化的每一条都是对生产数据的实测结论，与具体公司无关，因此属于引擎而非模型：

  · xlsx 声明的 <dimension> 常被写成 A1，不重置维度会把约 160 个文件读成"只有 1 行"
  · CSV 分隔符只看表头行判断——数据单元格里塞了制表符做防格式化，全文件统计会误判成 TSV
  · 制表符方向不一致：拼多多与千牛后置、抖店前导，必须双向去除
  · 重复列名按位置解析：千牛明细有 3 种签名存在重复列名，涉及 87 个文件，
    按列名建字典会静默丢数据
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
import tempfile
import zipfile
from pathlib import Path
from typing import IO, Any, Iterator

import openpyxl

try:
    from python_calamine import CalamineWorkbook

    _CALAMINE_OK = True
except ImportError:  # 没装也能跑，只是慢，且读不了 .xls
    CalamineWorkbook = None  # type: ignore[assignment]
    _CALAMINE_OK = False

from ..model.schema import ParseOptions
from .types import ControlTotal, FileRef, RawRow, RawTable

#: 只看表头行判断分隔符的候选集，按优先级。
_DELIMITERS = (",", "\t", ";", "|")

#: 表头行最长扫描长度。表头再长也不会超过这个量级，避免把整个大文件读进内存。
_HEADER_PROBE_BYTES = 1 << 20

#: 编码回退顺序。模板会指定编码，这里只在指定编码解不开时兜底。
_ENCODING_FALLBACKS = ("utf-8-sig", "gb18030", "utf-16", "utf-8")


class ParseError(Exception):
    """文件读不出来。消息必须说清是格式问题还是内容问题。"""


def digest(path: str | Path) -> str:
    """内容寻址。同一文件传两次不重复计算也不覆盖，靠的就是这个。"""
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def read_headers(path: str | Path, options: ParseOptions | None = None) -> list[list[str]]:
    """只读表头，用于识别阶段。xlsx 每个工作表一组表头。"""
    options = options or ParseOptions()
    path = Path(path)
    kind = _kind(path)
    if kind == "xlsx":
        return [hdr for _, hdr in _xlsx_headers(path, options)]
    return [_csv_headers(path, options)[0]]


def parse(
    path: str | Path,
    options: ParseOptions | None = None,
    sha: str | None = None,
    _depth: int = 0,
) -> list[RawTable]:
    """解析一个文件。xlsx 和压缩包都可能产出多张表。"""
    options = options or ParseOptions()
    path = Path(path)
    sha = sha or digest(path)
    kind = _kind(path)
    if kind == "xlsx":
        # xlsx 本身是 zip，先分清是表格还是装着表格的压缩包。
        if is_archive(path):
            if _depth >= _MAX_ARCHIVE_DEPTH:
                raise ParseError(f"{path.name} 压缩包套得太深，超过 {_MAX_ARCHIVE_DEPTH} 层不再往下拆")
            return _parse_archive(path, options, sha, _depth)
        return _parse_xlsx(path, options, sha)
    return [_parse_csv(path, options, sha)]


def _kind(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return "xlsx"
    if suffix in (".csv", ".txt", ".tsv"):
        return "csv"
    if suffix == "":
        # 内容寻址的快照文件没有扩展名，按魔数判断。
        with open(path, "rb") as fh:
            head = fh.read(4)
        return "xlsx" if head[:2] == b"PK" else "csv"
    if suffix == ".xls":
        raise ParseError(
            f"{path.name} 是旧版 xls 格式。请在 Excel 里另存为 xlsx 后重新上传。"
        )
    raise ParseError(f"{path.name} 的格式不支持（{suffix or '无扩展名'}）")


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #


#: 压缩包里最多再往下拆几层。防止套娃和 zip 炸弹。
_MAX_ARCHIVE_DEPTH = 2
#: xlsx 本身就是 zip，靠这个条目区分"是表格"还是"装着表格的压缩包"。
_XLSX_MARKER = "[Content_Types].xml"


def is_archive(path: Path) -> bool:
    """是不是装着表格的压缩包（而不是 xlsx 本身）。"""
    try:
        with zipfile.ZipFile(path) as zf:
            return _XLSX_MARKER not in zf.namelist()
    except (zipfile.BadZipFile, OSError):
        return False


def _entry_name(info: zipfile.ZipInfo) -> str:
    """还原压缩包里的中文文件名。

    zipfile 只在条目标了 UTF-8 标志位时按 UTF-8 解，否则一律按 cp437 解。
    国内平台导出的包基本都是 GBK 文件名，不还原就是一串乱码，
    连"这个包里装的是账务明细还是账务汇总"都看不出来。
    """
    if info.flag_bits & 0x800:
        return info.filename
    raw = info.filename.encode("cp437", errors="replace")
    for encoding in ("gbk", "utf-8"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return info.filename


def _parse_archive(path: Path, options: ParseOptions, sha: str, depth: int) -> list[RawTable]:
    """拆开压缩包，把里面每个条目当表格解析。

    实测支付宝账务明细就是按月打包下载的，一个包里两个 CSV（账务汇总 + 账务明细）。
    不拆包就等于整月对账数据不存在——而且是静默的，因为包本身"解析失败"看起来
    只是一个文件的问题。

    证据链保持完整：sha 仍是压缩包的，工作表名记成 `包内文件名`，
    下钻时能定位到"哪个包里的哪个文件的第几行"。
    """
    tables: list[RawTable] = []
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = _entry_name(info)
            suffix = Path(name).suffix.lower()
            if suffix not in (".csv", ".txt", ".xlsx", ".xlsm", ".xls", ".xlsb", ".zip"):
                continue
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp.write(zf.read(info))
                tmp_path = Path(tmp.name)
            try:
                inner = parse(tmp_path, options, sha=sha, _depth=depth + 1)
            except ParseError as exc:
                tables.append(
                    RawTable(
                        ref=FileRef(sha256=sha, filename=path.name, sheet=name),
                        headers=[],
                        notes=[f"压缩包里的 {name} 解析失败：{exc}"],
                    )
                )
                continue
            finally:
                tmp_path.unlink(missing_ok=True)
            for table in inner:
                # 工作表名前面挂上包内文件名，证据链才能一路追到具体条目。
                sheet = name if not table.ref.sheet else f"{name}#{table.ref.sheet}"
                table.ref = FileRef(sha256=sha, filename=path.name, sheet=sheet)
                tables.append(table)
    if not tables:
        raise ParseError(f"{path.name} 是压缩包，但里面没有能解析的表格文件")
    return tables


def _read_sheets(path: Path, options: ParseOptions) -> list[tuple[str, list[list]]]:
    """把工作簿读成 [(工作表名, 行列表)]。

    优先用 calamine：它是 Rust 实现，实测 64 MB 的对账表 2.3 秒读完，openpyxl 要
    15.1 秒，快 6.5 倍；而且同一个库能读 xlsx / xls / xlsb，openpyxl 只能读 xlsx，
    .xls 老二进制格式它根本打不开。

    openpyxl 留作兜底。两条路径读出来的值必须一致，read_sheets_agree 那个测试盯着这件事。
    """
    data = path.read_bytes()
    sheets: list[tuple[str, list[list]]] | None = None
    if _CALAMINE_OK:
        try:
            sheets = _read_with_calamine(data, options)
        except ParseError:
            raise
        except Exception:
            pass  # 落到 openpyxl 再试一次，读得出来就不算失败
    if sheets is None:
        sheets = _read_with_openpyxl(path, data, options)
    return [(name, _unmerge(data, name, rows)) for name, rows in sheets]


# --------------------------------------------------------------------------- #
# 合并单元格
# --------------------------------------------------------------------------- #

#: 一列里空格占比超过这个数，才值得去查它是不是合并出来的。
_SPARSE_SHARE = 0.2

_MERGE_REF = re.compile(rb'<mergeCell[^>]*\bref="([A-Z]+\d+:[A-Z]+\d+)"')


def _unmerge(data: bytes, sheet: str, rows: list[list]) -> list[list]:
    """把合并单元格的值填回区域内的每一行。

    合并单元格的语义是「这个值属于整个区域」，但 xlsx 只把值存在左上角那一格，
    其余格子是空的。读的时候不还原，下游看到的就是一列半空的数据。

    实测 1688 那张订单明细：一个订单有几行商品，订单号和订单创建时间都是合并的，
    3,086 行里只有 1,516 行有值。不还原的话一半的行会因为订单号为空被当成合计行丢掉，
    剩下的因为没有日期落不进账期。制表的人是知道这个坑的——他手工把订单号那一列
    展开填充后另存了一份贴在旁边。这件事该引擎做，不该靠人记得做、也不该靠模板
    逐列去指哪一列要向下填。

    合并区域在 xlsx 里是有明确记录的，读出来照着填就行，不用猜。
    只是那段记录写在工作表 XML 的末尾，要拿到它就得把整张表解压一遍——
    30 万行的运费表这么干太亏。所以先看已经读出来的行有没有合并的痕迹
    （某列大量空格，且每个空格上方都有值），有痕迹才去查真相。
    """
    if not rows or not _may_have_merges(rows):
        return rows
    ranges = _merged_ranges(data, sheet)
    for r1, c1, r2, c2 in ranges:
        if r1 >= len(rows) or c1 >= len(rows[r1]):
            continue
        value = rows[r1][c1]
        if value is None or value == "":
            continue
        for r in range(r1 + 1, min(r2 + 1, len(rows))):
            for c in range(c1, min(c2 + 1, len(rows[r]))):
                if rows[r][c] is None or rows[r][c] == "":
                    rows[r][c] = value
    return rows


def _may_have_merges(rows: list[list]) -> bool:
    """看行里有没有合并的痕迹：某一列大量空格，且空格上方都有值。

    纯粹是为了省掉不必要的解压。判断错了只会多读或少读一次 XML，
    真正填什么以 XML 里记的合并区域为准。
    """
    width = max((len(r) for r in rows[:200]), default=0)
    for c in range(width):
        column = [r[c] if c < len(r) else None for r in rows]
        blank = sum(1 for v in column if v is None or v == "")
        if blank < len(column) * _SPARSE_SHARE:
            continue
        # 合并区域的空格必然紧跟在有值的格子下面，不会出现在整列开头。
        if any(v not in (None, "") for v in column[:20]):
            return True
    return False


def _merged_ranges(data: bytes, sheet: str) -> list[tuple[int, int, int, int]]:
    """从 xlsx 里读某张表的合并区域，返回 0 基的 (起行, 起列, 止行, 止列)。"""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            target = _sheet_xml_path(zf, sheet)
            if target is None:
                return []
            with zf.open(target) as fh:
                blob = _tail_after(fh, b"<mergeCells")
        return [_ref_to_box(m.group(1).decode()) for m in _MERGE_REF.finditer(blob)]
    except (zipfile.BadZipFile, KeyError, ValueError):
        return []


def _tail_after(fh, marker: bytes, chunk: int = 1 << 20) -> bytes:
    """流式读到标记出现为止，只留标记之后的内容。

    合并区域记在 sheetData 后面，也就是文件末尾。整张表读进内存的话，
    大表会吃掉几百兆，所以边解压边丢。
    """
    tail = b""
    found = False
    while True:
        block = fh.read(chunk)
        if not block:
            return tail if found else b""
        if found:
            tail += block
            continue
        buf = tail[-len(marker):] + block
        i = buf.find(marker)
        if i >= 0:
            found = True
            tail = buf[i:]
        else:
            tail = block[-len(marker):]


def _sheet_xml_path(zf: zipfile.ZipFile, sheet: str) -> str | None:
    book = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
    m = re.search(rf'<sheet[^>]*\bname="{re.escape(sheet)}"[^>]*/>', book)
    if not m:
        return None
    rid = re.search(r'r:id="([^"]+)"', m.group(0))
    if not rid:
        return None
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    rel = re.search(rf'<Relationship[^>]*\bId="{re.escape(rid.group(1))}"[^>]*/>', rels)
    if not rel:
        return None
    tgt = re.search(r'Target="([^"]+)"', rel.group(0))
    if not tgt:
        return None
    path = tgt.group(1).lstrip("/")
    return path if path.startswith("xl/") else f"xl/{path}"


def _ref_to_box(ref: str) -> tuple[int, int, int, int]:
    """A2:C5 → (1, 0, 4, 2)。"""
    start, _, end = ref.partition(":")
    r1, c1 = _cell_to_rc(start)
    r2, c2 = _cell_to_rc(end or start)
    return r1, c1, r2, c2


def _cell_to_rc(cell: str) -> tuple[int, int]:
    m = re.match(r"([A-Z]+)(\d+)", cell)
    if not m:
        raise ValueError(cell)
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)) - 1, col - 1


def _read_with_calamine(data: bytes, options: ParseOptions) -> list[tuple[str, list[list]]]:
    wb = CalamineWorkbook.from_filelike(io.BytesIO(data))
    names = wb.sheet_names
    if options.sheet:
        if options.sheet not in names:
            raise ParseError(f"工作表 {options.sheet} 不存在，实际有：{', '.join(names)}")
        names = [options.sheet]
    return [(name, wb.get_sheet_by_name(name).to_python()) for name in names]


def _read_with_openpyxl(path: Path, data: bytes, options: ParseOptions) -> list[tuple[str, list[list]]]:
    try:
        if path.suffix.lower() in (".xlsx", ".xlsm"):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except zipfile.BadZipFile as exc:
        raise ParseError(f"{path.name} 不是有效的 xlsx（zip 结构损坏）") from exc
    except Exception as exc:  # openpyxl 抛的异常类型很杂
        raise ParseError(f"{path.name} 无法作为 xlsx 打开：{exc}") from exc
    try:
        sheets = _sheets(wb, options)
        out = []
        for ws in sheets:
            # 关键：不重置维度，声明为 A1 的文件会被读成只有 1 行。
            ws.reset_dimensions = options.reset_xlsx_dimension
            out.append((ws.title, [list(r) for r in ws.iter_rows(values_only=True)]))
        return out
    finally:
        wb.close()


def _sheets(wb, options: ParseOptions):
    if options.sheet:
        if options.sheet not in wb.sheetnames:
            raise ParseError(f"工作表 {options.sheet} 不存在，实际有：{', '.join(wb.sheetnames)}")
        return [wb[options.sheet]]
    return list(wb.worksheets)


def _xlsx_headers(path: Path, options: ParseOptions) -> Iterator[tuple[str, list[str]]]:
    for name, rows in _read_sheets(path, options):
        if len(rows) > options.header_row:
            yield name, _clean_header(rows[options.header_row], options)
        else:
            yield name, []


def _parse_xlsx(path: Path, options: ParseOptions, sha: str) -> list[RawTable]:
    tables: list[RawTable] = []
    for name, values_rows in _read_sheets(path, options):
        ref = FileRef(sha256=sha, filename=path.name, sheet=name)
        headers: list[str] = []
        rows: list[RawRow] = []
        data_starts = options.header_row + 1 + options.skip_after_header
        for i, values in enumerate(values_rows):
            if i == options.header_row:
                headers = _clean_header(values, options)
                continue
            if i < data_starts or not headers:
                continue
            cells = _clean_cells(values, len(headers), options)
            if cells is None:
                continue
            rows.append(RawRow(row_no=i + 1, cells=cells))
        table = RawTable(ref=ref, headers=headers, rows=rows)
        if not headers:
            table.notes.append(f"工作表 {name} 第 {options.header_row + 1} 行没有表头")
        tables.append(table)
    return tables


# --------------------------------------------------------------------------- #
# csv
# --------------------------------------------------------------------------- #


def _decode(path: Path, options: ParseOptions) -> tuple[str, str]:
    """按模板指定编码读，解不开才回退。返回 (文本, 实际编码)。"""
    data = path.read_bytes()
    order = (options.encoding, *(e for e in _ENCODING_FALLBACKS if e != options.encoding))
    for enc in order:
        try:
            return data.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode(options.encoding, errors="replace"), options.encoding + "(有损)"


#: 注释行前缀。支付宝导出的账务明细首尾各有若干行以此开头的元信息。
_COMMENT_PREFIX = "#"

#: 找表头最多往下看多少行。再多说明这个文件本来就不是表格。
_HEADER_SEARCH_ROWS = 30

#: 从注释行里认控制总数。实测格式：`#支出合计：75171笔，共-540182.61元`
_CONTROL_PATTERN = re.compile(
    r"^#\s*(?P<label>[^：:]{2,12})\s*[：:]\s*"
    r"(?:(?P<count>-?[\d,]+)\s*笔\s*[，,]?\s*)?"
    r"(?:共\s*)?(?P<amount>-?[\d,]+\.?\d*)\s*元"
)

#: 控制总数的标签对应看哪个方向的金额。
_CONTROL_DIRECTION = {"支出合计": "outgo", "收入合计": "income"}


def _locate_header(lines: list[str], options: ParseOptions) -> int:
    """找表头在第几行。

    显式声明了就用声明的。没声明时跳过开头的注释行——支付宝导出的账务明细前面有
    4 行 `#` 元信息（账号、起止日期、分隔线），表头在第 5 行。按第 1 行当表头的话
    整个文件会被读成 1 列，全量语料里有 113 个文件栽在这上面。
    """
    if options.header_row:
        return options.header_row
    for i, line in enumerate(lines[:_HEADER_SEARCH_ROWS]):
        if line.strip() and not line.startswith(_COMMENT_PREFIX):
            return i
    return 0


def _extract_controls(lines: list[str]) -> list[ControlTotal]:
    """把注释行里的控制总数抠出来。

    文件自己声明了有多少笔、合计多少钱，解析完拿它对一遍就知道有没有读丢。
    这是文件白给的正确性证据。
    """
    out: list[ControlTotal] = []
    for line in lines:
        if not line.startswith(_COMMENT_PREFIX):
            continue
        m = _CONTROL_PATTERN.match(line.strip())
        if not m:
            continue
        label = m.group("label").strip()
        count = m.group("count")
        amount = m.group("amount")
        out.append(
            ControlTotal(
                label=label,
                count=int(count.replace(",", "")) if count else None,
                amount=float(amount.replace(",", "")) if amount else None,
                direction=_CONTROL_DIRECTION.get(label, "both"),
                raw=line.strip(),
            )
        )
    return out


def _csv_headers(path: Path, options: ParseOptions) -> tuple[list[str], str, str, int]:
    text, enc = _decode(path, options)
    lines = text.splitlines()
    header_row = _locate_header(lines, options)
    if header_row >= len(lines):
        raise ParseError(f"{path.name} 只有 {len(lines)} 行，取不到第 {header_row + 1} 行做表头")
    header_line = lines[header_row][:_HEADER_PROBE_BYTES]
    delimiter = options.delimiter or _sniff_delimiter(header_line)
    headers = _clean_header(next(csv.reader([header_line], delimiter=delimiter)), options)
    return headers, delimiter, enc, header_row


def _sniff_delimiter(header_line: str) -> str:
    """只看表头行。数据单元格里塞制表符做防格式化，全文件统计会误判成 TSV。"""
    counts = {d: header_line.count(d) for d in _DELIMITERS}
    best = max(counts, key=lambda d: counts[d])
    return best if counts[best] else ","


def _parse_csv(path: Path, options: ParseOptions, sha: str) -> RawTable:
    headers, delimiter, enc, header_row = _csv_headers(path, options)
    text, _ = _decode(path, options)
    lines = text.splitlines()
    ref = FileRef(sha256=sha, filename=path.name, sheet=None)
    table = RawTable(ref=ref, headers=headers, controls=_extract_controls(lines))
    if enc != options.encoding:
        table.notes.append(f"编码按 {options.encoding} 解不开，实际用 {enc}")
    if not options.delimiter and delimiter != ",":
        table.notes.append(f"表头行判定分隔符为 {delimiter!r}")
    if header_row and not options.header_row:
        table.notes.append(f"表头在第 {header_row + 1} 行，前面 {header_row} 行是 # 开头的元信息")

    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    data_starts = header_row + 1 + options.skip_after_header
    comment_rows = 0
    for i, values in enumerate(reader):
        if i < data_starts:
            continue
        # 尾部同样有 # 开头的注释行（列表结束标记、合计、导出时间），
        # 不剔掉会变成几行只有一列的垃圾数据混进金额统计。
        if values and str(values[0]).startswith(_COMMENT_PREFIX):
            comment_rows += 1
            continue
        cells = _clean_cells(values, len(headers), options)
        if cells is None:
            continue
        table.rows.append(RawRow(row_no=i + 1, cells=cells))
    if comment_rows:
        table.notes.append(f"跳过 {comment_rows} 行 # 开头的注释行")
    if table.controls:
        table.notes.append(
            "文件自带控制总数：" + "；".join(c.raw.lstrip("#") for c in table.controls)
        )
    return table


# --------------------------------------------------------------------------- #
# 单元格清洗
# --------------------------------------------------------------------------- #


def _clean_header(values: Any, options: ParseOptions) -> list[str]:
    out = [_strip(v, options) for v in (values or ())]
    while out and not out[-1]:
        out.pop()
    return out


def _clean_cells(values: Any, width: int, options: ParseOptions) -> tuple[Any, ...] | None:
    """返回 None 表示整行为空，跳过。宽度不足补空，超出的保留。

    超出表头宽度的单元格不能丢：这批文件第一行是人手写的说明文字，只占稀疏几个
    单元格，真表头在第二行。按第一行的宽度截断会把真表头一起截没，识别就永远
    发现不了它。归一时只按角色绑定的列取值，多出来的列留着无害。
    """
    cells = [_strip(v, options) if isinstance(v, str) or v is None else v for v in (values or ())]
    if not any(c not in (None, "") for c in cells):
        return None
    if len(cells) < width:
        cells.extend([None] * (width - len(cells)))
    return tuple(cells)


def _strip(value: Any, options: ParseOptions) -> str:
    """双向去除空白与制表符。方向不一致，只去一边会留下脏字符。"""
    if value is None:
        return ""
    s = str(value)
    if options.strip_tabs:
        s = s.strip("\t\u3000 \r\n\ufeff")
    else:
        s = s.strip()
    return "" if s in options.null_tokens else s


def open_text(path: str | Path, options: ParseOptions | None = None) -> IO[str]:
    """按模板编码打开文本文件。给需要逐行处理的调用方。"""
    options = options or ParseOptions()
    text, _ = _decode(Path(path), options)
    return io.StringIO(text)
