"""资产层：科目字典与运营归属的导入。

现有的计算资产是这个项目最被低估的财富。`运营链接.xlsx` 里有四张人手维护的映射表，
结构是 `业务描述 → 业务小类 → 业务大类`，覆盖阿里、拼多多、抖店、京东1688。
这批数据是"七套口径统一为一套"的现成依据，不用重新整理。

导入器只执行 `asset-import.yaml` 里记录的归一决策，决策本身可追溯。
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import yaml


@dataclass
class DictRow:
    platform: str
    raw: str
    minor: str
    major: str
    naturally_unlinked: bool
    #: 归一前的原始大类写法，用于审计。
    major_source: str = ""
    #: 这一条是主名还是别名（阿里的千牛明细叫法）。
    is_alias: bool = False


@dataclass
class ImportReport:
    """导入结果。未能归一的必须逐条列出原因，不允许静默跳过。"""

    rows: list[DictRow] = field(default_factory=list)
    responsibility: list[tuple[str, str, str, str]] = field(default_factory=list)
    #: 大类归一表里没有的写法。
    unmapped_majors: dict[str, int] = field(default_factory=dict)
    skipped: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def summary(self) -> str:
        alias = sum(1 for r in self.rows if r.is_alias)
        lines = [
            f"科目字典 {len(self.rows)} 条（其中别名 {alias} 条）",
            f"运营归属 {len(self.responsibility)} 条",
        ]
        if self.unmapped_majors:
            lines.append(
                "未归一的大类写法 "
                + "、".join(f"{k}({v})" for k, v in sorted(self.unmapped_majors.items()))
            )
        if self.skipped:
            lines.append(f"跳过 {len(self.skipped)} 条：" + "；".join(self.skipped[:3]))
        return "\n".join(lines)


def load_rules(path: str | Path) -> dict[str, Any]:
    return yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}


def import_assets(workbook_path: str | Path, rules: dict[str, Any]) -> ImportReport:
    """从运营链接工作簿导入字典与运营归属。"""
    report = ImportReport()
    wb = _open(workbook_path)
    try:
        canonical: dict[str, str] = rules.get("major_canonical", {}) or {}
        non_profit = set(rules.get("non_profit_majors", []) or [])

        for spec in rules.get("dictionary_sheets", []) or []:
            _import_dictionary_sheet(wb, spec, canonical, non_profit, report)

        for sheet in rules.get("responsibility_sheets", []) or []:
            _import_responsibility_sheet(wb, str(sheet), rules, report)
    finally:
        wb.close()

    _dedupe(report)
    return report


def _open(path: str | Path):
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    # 内容寻址的快照文件没有扩展名。
    return openpyxl.load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)


def _rows(wb, sheet: str) -> tuple[list[str], list[tuple]]:
    if sheet not in wb.sheetnames:
        return [], []
    ws = wb[sheet]
    ws.reset_dimensions = True
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return [], []
    headers = [str(c or "").strip() for c in data[0]]
    body = [r for r in data[1:] if any(c not in (None, "") for c in r)]
    return headers, body


def _cell(row: tuple, headers: list[str], name: str | None) -> str:
    if not name or name not in headers:
        return ""
    i = headers.index(name)
    return str(row[i]).strip() if i < len(row) and row[i] is not None else ""


def _import_dictionary_sheet(
    wb,
    spec: dict[str, Any],
    canonical: dict[str, str],
    non_profit: set[str],
    report: ImportReport,
) -> None:
    sheet = spec["sheet"]
    headers, body = _rows(wb, sheet)
    if not headers:
        report.skipped.append(f"工作表 {sheet} 不存在或为空")
        return

    platform = spec["platform"]
    unlinked_values = {str(v) for v in spec.get("unlinked_values", []) or []}

    for row in body:
        raw = _cell(row, headers, spec["raw"])
        if not raw:
            continue
        major_source = _cell(row, headers, spec["major"])
        major = canonical.get(major_source)
        if major is None:
            report.unmapped_majors[major_source or "(空)"] = (
                report.unmapped_majors.get(major_source or "(空)", 0) + 1
            )
            report.skipped.append(f"{sheet}·{raw}：大类 {major_source or '(空)'} 无归一规则")
            continue

        # 抖店那张表业务小类整列为空，回退到业务描述。
        minor = _cell(row, headers, spec.get("minor")) or raw
        flagged = _cell(row, headers, spec.get("unlinked_flag")) in unlinked_values
        unlinked = flagged or major in non_profit

        scope = _cell(row, headers, spec.get("scope"))
        plat = _scoped_platform(platform, scope)

        report.rows.append(
            DictRow(
                platform=plat, raw=raw, minor=minor, major=major,
                naturally_unlinked=unlinked, major_source=major_source,
            )
        )

        alias = _cell(row, headers, spec.get("alias"))
        if alias and alias != raw:
            report.rows.append(
                DictRow(
                    platform=plat, raw=alias, minor=minor, major=major,
                    naturally_unlinked=unlinked, major_source=major_source, is_alias=True,
                )
            )


def _scoped_platform(platform: str, scope: str) -> str:
    """京东和 1688 在同一张表里，靠店铺列区分。"""
    if platform != "jd_1688" or not scope:
        return platform
    return {"京东": "jd", "1688": "alibaba1688"}.get(scope, platform)


_MONTH = re.compile(r"^(\d{4})姓名$")


def _import_responsibility_sheet(wb, sheet: str, rules: dict[str, Any], report: ImportReport) -> None:
    """把按月宽表转成长表。宽表每加一个月要加一列，长表不用改结构。"""
    headers, body = _rows(wb, sheet)
    if not headers:
        report.skipped.append(f"工作表 {sheet} 不存在或为空")
        return

    key_col = rules.get("responsibility_key", "宝贝编码")
    store_col = rules.get("responsibility_store", "")
    pattern = re.compile(rules.get("responsibility_column_pattern", _MONTH.pattern))

    months = [(i, h) for i, h in enumerate(headers) if pattern.match(h)]
    if not months:
        report.notes.append(f"工作表 {sheet} 没有按月姓名列")
        return

    for row in body:
        product = _cell(row, headers, key_col)
        if not product:
            continue
        store = _cell(row, headers, store_col)
        for i, header in months:
            owner = str(row[i]).strip() if i < len(row) and row[i] is not None else ""
            if not owner:
                continue
            yymm = pattern.match(header).group(1)  # type: ignore[union-attr]
            period = f"20{yymm[:2]}-{yymm[2:]}"
            report.responsibility.append((product, period, owner, store))
    report.notes.append(f"工作表 {sheet} 展开 {len(months)} 个月")


def _dedupe(report: ImportReport) -> None:
    """同一 (平台, 科目) 只保留一条。主名优先于别名，后出现的覆盖前面的。"""
    seen: dict[tuple[str, str], int] = {}
    kept: list[DictRow] = []
    conflicts = 0
    for row in report.rows:
        key = (row.platform, row.raw)
        if key in seen:
            prior = kept[seen[key]]
            if prior.major != row.major:
                conflicts += 1
                report.notes.append(
                    f"{row.platform}·{row.raw} 归类冲突：{prior.major} 与 {row.major}，取后者"
                )
            kept[seen[key]] = row
            continue
        seen[key] = len(kept)
        kept.append(row)
    report.rows = kept
    if conflicts:
        report.notes.append(f"共 {conflicts} 处归类冲突")

    resp = {(p, m): (o, s) for p, m, o, s in report.responsibility}
    report.responsibility = [(p, m, o, s) for (p, m), (o, s) in sorted(resp.items())]


# --------------------------------------------------------------------------- #
# 落盘
# --------------------------------------------------------------------------- #


def write_dictionary(report: ImportReport, path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["platform", "raw", "minor", "major", "naturally_unlinked"])
        for row in sorted(report.rows, key=lambda r: (r.platform, r.major, r.raw)):
            w.writerow([row.platform, row.raw, row.minor, row.major, int(row.naturally_unlinked)])


def write_responsibility(report: ImportReport, path: str | Path) -> None:
    """商品归属有 62 万行，写成 gzip。它是派生数据，随时可以从工作簿重新导出。"""
    import gzip

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    opener = gzip.open if path.suffix == ".gz" else open
    with opener(path, "wt", encoding="utf-8", newline="") as fh:  # type: ignore[operator]
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["product_id", "period", "owner", "store"])
        w.writerows(report.responsibility)
