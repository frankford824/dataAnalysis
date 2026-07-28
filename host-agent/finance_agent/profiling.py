from __future__ import annotations

import csv
import hashlib
import json
import re
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import openpyxl
import polars as pl

from .models import ColumnProfile, StructureProfile

HEADER_ALIASES: dict[str, frozenset[str]] = {
    "order_id": frozenset(
        {"订单号", "订单编号", "主订单编号", "子订单编号", "交易订单号", "order_id"}
    ),
    "business_date": frozenset(
        {"日期", "业务日期", "下单时间", "付款时间", "创建时间", "交易时间", "date"}
    ),
    "sales": frozenset(
        {"销售额", "实付金额", "买家实付金额", "成交金额", "收入", "sales"}
    ),
    "refund": frozenset({"退款", "退款金额", "退货退款", "refund"}),
    "fee": frozenset(
        {"费用", "平台费用", "手续费", "广告消耗", "广告费", "运费", "fee"}
    ),
    "cost": frozenset({"成本", "商品成本", "采购成本", "cost"}),
    "settlement": frozenset(
        {"收支金额", "收入金额", "支出金额", "账务流水", "支付宝流水号", "settlement"}
    ),
}

CLASSIFICATION_RULES: tuple[tuple[str, frozenset[str]], ...] = (
    ("orders", frozenset({"order_id", "business_date", "sales"})),
    ("advertising", frozenset({"business_date", "fee"})),
    ("product_cost", frozenset({"order_id", "cost"})),
    ("settlement", frozenset({"business_date", "settlement"})),
    ("shipping", frozenset({"business_date", "fee"})),
)


class UnsupportedProfileError(ValueError):
    pass


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s\u3000_\-（）()【】\[\]:：/]+", "", text).casefold()


def canonical_header(value: Any) -> str:
    normalized = normalize_header(value)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in {normalize_header(alias) for alias in aliases}:
            return canonical
    return normalized


def schema_fingerprint(headers: Iterable[Any], inferred_types: Iterable[str]) -> str:
    payload = {
        "headers": [canonical_header(header) for header in headers],
        "types": list(inferred_types),
        "version": 1,
    }
    encoded = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def classify(headers: Iterable[Any], purpose_hint: str | None = None) -> tuple[str, str]:
    canonical = {canonical_header(header) for header in headers}
    candidates: list[str] = []
    for name, required in CLASSIFICATION_RULES:
        if required.issubset(canonical):
            candidates.append(name)
    if len(candidates) == 1:
        return candidates[0], "high"
    if purpose_hint in {
        "orders",
        "advertising",
        "settlement",
        "product_cost",
        "shipping",
    }:
        return str(purpose_hint), "medium"
    if len(candidates) > 1:
        return "needs_confirmation", "ambiguous"
    return "unknown", "low"


def profile_file(
    path: Path,
    purpose_hint: str | None = None,
    sample_rows: int = 200,
) -> StructureProfile:
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        return _profile_csv(path, purpose_hint, sample_rows)
    if suffix in {".xlsx", ".xlsm"}:
        return _profile_xlsx(path, purpose_hint, sample_rows)
    if suffix == ".xls":
        raise UnsupportedProfileError("旧版 .xls 暂不自动读取，请另存为 .xlsx 或 CSV")
    if suffix == ".pbix":
        raise UnsupportedProfileError("PBIX 只登记元数据，不参与运行时金额解析")
    raise UnsupportedProfileError(f"不支持结构画像的文件类型: {suffix}")


def _profile_csv(
    path: Path, purpose_hint: str | None, sample_rows: int
) -> StructureProfile:
    encoding = _detect_encoding(path)
    separator = _detect_separator(path, encoding)
    frame = pl.read_csv(
        path,
        encoding="utf8" if encoding.startswith("utf-8") else "utf8-lossy",
        separator=separator,
        n_rows=sample_rows,
        infer_schema_length=min(sample_rows, 100),
        ignore_errors=True,
        try_parse_dates=False,
    )
    return _profile_frame(
        frame,
        str(path),
        "csv",
        None,
        1,
        purpose_hint,
        warnings=() if encoding.startswith("utf-8") else (f"源编码为 {encoding}",),
    )


def _profile_xlsx(
    path: Path, purpose_hint: str | None, sample_rows: int
) -> StructureProfile:
    workbook = openpyxl.load_workbook(
        path, read_only=True, data_only=True, keep_links=False
    )
    try:
        visible = [
            sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"
        ]
        if not visible:
            raise UnsupportedProfileError("工作簿没有可见 Sheet")
        sheet = visible[0]
        rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
        header_index = _find_header_row(rows)
        headers = [str(value or "").strip() for value in rows[header_index]]
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            raise UnsupportedProfileError("未找到有效表头")
        data: list[list[Any]] = []
        for row in sheet.iter_rows(
            min_row=header_index + 2,
            max_row=header_index + 1 + sample_rows,
            max_col=len(headers),
            values_only=True,
        ):
            data.append(list(row))
        columns: dict[str, list[Any]] = {}
        seen: dict[str, int] = {}
        for index, header in enumerate(headers):
            base = header or f"column_{index + 1}"
            count = seen.get(base, 0) + 1
            seen[base] = count
            name = base if count == 1 else f"{base}_{count}"
            columns[name] = [row[index] if index < len(row) else None for row in data]
        frame = pl.DataFrame(columns, strict=False)
        return _profile_frame(
            frame,
            str(path),
            "xlsx",
            sheet.title,
            header_index + 1,
            purpose_hint,
        )
    finally:
        workbook.close()


def _profile_frame(
    frame: pl.DataFrame,
    path: str,
    file_type: str,
    sheet: str | None,
    header_row: int,
    purpose_hint: str | None,
    warnings: tuple[str, ...] = (),
) -> StructureProfile:
    columns: list[ColumnProfile] = []
    for name, dtype in frame.schema.items():
        series = frame[name]
        samples = tuple(
            _masked_sample(value)
            for value in series.drop_nulls().head(3).to_list()
            if str(value).strip()
        )
        columns.append(
            ColumnProfile(
                name=name,
                inferred_type=str(dtype),
                non_null_count=int(series.is_not_null().sum()),
                sample_values=samples,
            )
        )
    classification, confidence = classify(frame.columns, purpose_hint)
    return StructureProfile(
        path=path,
        file_type=file_type,
        sheet=sheet,
        header_row=header_row,
        row_count_sampled=frame.height,
        columns=tuple(columns),
        fingerprint=schema_fingerprint(
            frame.columns, (str(dtype) for dtype in frame.dtypes)
        ),
        classification=classification,
        classification_confidence=confidence,
        warnings=warnings,
    )


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    scored: list[tuple[int, int]] = []
    known = {
        normalize_header(alias)
        for aliases in HEADER_ALIASES.values()
        for alias in aliases
    }
    for index, row in enumerate(rows):
        nonempty = [value for value in row if value is not None and str(value).strip()]
        alias_hits = sum(normalize_header(value) in known for value in nonempty)
        score = alias_hits * 10 + min(len(nonempty), 9)
        scored.append((score, -index))
    if not scored or max(scored)[0] < 2:
        raise UnsupportedProfileError("前 20 行未找到可信表头")
    return -max(scored)[1]


def _detect_encoding(path: Path) -> str:
    sample = path.read_bytes()[:65536]
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise UnsupportedProfileError("CSV 编码无法识别")


def _detect_separator(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(8192)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        return ","


def _masked_sample(value: Any) -> str:
    text = str(value).strip()
    digest = hashlib.sha256(text.encode("utf-8")).hexdigest()[:12]
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text.replace(",", "")):
        scale = len(text.rsplit(".", 1)[1]) if "." in text else 0
        return f"<number:scale={scale}:hash={digest}>"
    if re.fullmatch(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}.*", text):
        month = re.split(r"[-/.]", text)[:2]
        return f"<date:{month[0]}-{int(month[1]):02d}:hash={digest}>"
    return f"<text:length={len(text)}:hash={digest}>"
